"""
Generar Excel de relevamiento de concesionarias — Valle de Punilla
Paulero Studio · Planilla de prospección comercial
"""
import sys, os
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from templates.base import (
    FONT_NAME, PRIMARY, PRIMARY_LIGHT, NEUTRAL_900, NEUTRAL_600,
    NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    fill_header, font_header, font_body, align_text, align_header,
    setup_sheet, auto_fit_columns,
)

# ============================================================
# DATA: Concesionarias relevadas
# ============================================================
# Columns: nombre, ubicacion, direccion, telefono, tipo, url_web, redes, estado_web, prioridad, notas

DEALERSHIPS = [
    # ===== VILLA CARLOS PAZ =====
    {
        "nombre": "Automotores Martínez",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Villa Carlos Paz (sin dirección pública clara)",
        "telefono": "(3541) 206969 · (3541) 760004",
        "tipo": "Autos 0km + usados multimarca",
        "url_web": "https://sites.google.com/view/automotoresmartinez",
        "redes": "FB: Automotores Martínez",
        "estado_web": "Amateur (Google Sites)",
        "prioridad": "Alta",
        "notas": "0km multimarca. Solo Google Sites muy básico. Oportunidad clara de Landing o Sitio Completo con catálogo.",
    },
    {
        "nombre": "AutoFamily",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. Perón 711, Villa Carlos Paz",
        "telefono": "(03541) 760004",
        "tipo": "Autos usados y 0km multimarca",
        "url_web": "https://autofamily.com.ar",
        "redes": "IG: @autofamily__ · +30 años en rubro",
        "estado_web": "Existe (no accesible a bots)",
        "prioridad": "Media",
        "notas": "+30 años trayectoria, +100 autos en stock. Sitio web registró timeout — verificar manualmente. Podría necesitar refresh.",
    },
    {
        "nombre": "DG Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. San Martín 1678 · Cárcano 696, Villa Carlos Paz",
        "telefono": "Desde IG (no público en web)",
        "tipo": "Autos usados + 0km multimarca",
        "url_web": "https://www.dgautomotores.com.ar",
        "redes": "IG activo",
        "estado_web": "Decente pero simple",
        "prioridad": "Baja",
        "notas": "Tiene sitio pero HTML muy liviano (2KB) — probablemente landing estática. Sin H1. Oportunidad de e-commerce con catálogo.",
    },
    {
        "nombre": "Montironi Ford",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. Illia 615, Villa Carlos Paz",
        "telefono": "0800-444-1111",
        "tipo": "0km oficial Ford",
        "url_web": "https://montironiford.com",
        "redes": "IG: @montironiok",
        "estado_web": "Decente (WordPress corporativo)",
        "prioridad": "Baja",
        "notas": "Concesionaria oficial Ford. Sitio WordPress corporativo, probablemente con soporte de marca. Difícil venderle web pero podría venderle chatbot IA.",
    },
    {
        "nombre": "CBA Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. Cárcano 1909, Villa Carlos Paz",
        "telefono": "(3541) 377489 · (3541) 232400",
        "tipo": "0km + usados + consignaciones",
        "url_web": "https://cbaautomotores.com.ar",
        "redes": "IG: @cba_automotores.vcp",
        "estado_web": "Decente",
        "prioridad": "Baja",
        "notas": "Sitio bien armado (128KB HTML, viewport + meta). Financiación hasta 75% solo DNI. Pitch enfocado en chatbot IA para calificar leads.",
    },
    {
        "nombre": "Angle Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Fleming S/N, La Cuesta, Villa Carlos Paz",
        "telefono": "(03541) 43-4847",
        "tipo": "Autos usados multimarca",
        "url_web": "Sin web propia",
        "redes": "FB: Angle Automotores",
        "estado_web": "Sin web (solo Facebook + Autocosmos)",
        "prioridad": "Alta",
        "notas": "Solo listado en Facebook y Autocosmos. Empresa joven pero con experiencia. Oportunidad clara para Landing Page con stock.",
    },
    {
        "nombre": "Torino Select Garage",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Ramón J. Cárcano 1608 esq. Ibsen, Villa Carlos Paz",
        "telefono": "WhatsApp 351 242 9960",
        "tipo": "Autos usados premium",
        "url_web": "Sin web propia",
        "redes": "IG: @torinoselectgarage",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Usados premium/gama alta. Solo Instagram. Tiene dirección física y WhatsApp. Candidato ideal para Landing con galería de autos premium.",
    },
    {
        "nombre": "Redolfi Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Diego de Velázquez 40, Villa Carlos Paz",
        "telefono": "+54 9 3541 279366",
        "tipo": "Autos usados",
        "url_web": "Sin web propia",
        "redes": "FB: Redolfi Automotores",
        "estado_web": "Sin web (solo Facebook)",
        "prioridad": "Alta",
        "notas": "Solo Facebook. Dirección y WhatsApp públicos. Pitch directo para Landing Page 250 USD.",
    },
    {
        "nombre": "V-Cars Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Villa Carlos Paz (dirección no clara)",
        "telefono": "Desde IG",
        "tipo": "0km Ford, Fiat, Chevrolet, Renault, VW",
        "url_web": "Sin web propia",
        "redes": "IG mencionado",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Cumplió 1 año en Carlos Paz. Solo IG. Menciona 5 marcas de 0km. Buen candidato para Sitio Completo con catálogo 0km.",
    },
    {
        "nombre": "Aspen Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Villa Carlos Paz",
        "telefono": "Desde IG",
        "tipo": "Camionetas 0km (Chevrolet, VW, Toyota, Ford)",
        "url_web": "Sin web propia",
        "redes": "IG mencionado",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Nicho interesante: solo camionetas. Sin web. E-commerce con filtros por marca/modelo sería ideal para catálogo.",
    },
    {
        "nombre": "Pablo Automotores",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Rene Simón 1380, Villa Carlos Paz",
        "telefono": "Desde FB",
        "tipo": "VW 0km y usados especialista",
        "url_web": "Sin web propia",
        "redes": "IG + FB",
        "estado_web": "Sin web (solo IG + FB)",
        "prioridad": "Alta",
        "notas": "Especialista VW. Sin web propia. Nicho claro — Landing enfocada en VW.",
    },
    {
        "nombre": "Giorgis Motoworld",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. Libertad 21, Villa Carlos Paz",
        "telefono": "03541-420994 / 427529",
        "tipo": "Motos Honda oficial",
        "url_web": "Sin web propia",
        "redes": "FB: Giorgis Motoworld",
        "estado_web": "Sin web (solo Facebook + Honda oficial)",
        "prioridad": "Alta",
        "notas": "Concesionario oficial Honda. Sin web propia (solo fanpage y listing Honda). 3 años garantía oficial. Oportunidad alta — la marca debería exigirle mejor web.",
    },
    {
        "nombre": "AMES Motos (sucursal Carlos Paz)",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. Libertad 365, Villa Carlos Paz",
        "telefono": "(03541) 439180 · WhatsApp (0351) 153111944",
        "tipo": "Motos Honda oficial",
        "url_web": "Sin web propia",
        "redes": "IG: @ames.motos",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Honda oficial. 40 años trayectoria. Sucursal también en Córdoba Capital. Sin web propia. Candidato fuerte para Sitio Completo con catálogo Honda.",
    },
    {
        "nombre": "E-Motors VCP",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Villa Carlos Paz",
        "telefono": "Desde IG",
        "tipo": "Motos eléctricas + repuestos",
        "url_web": "Sin web propia",
        "redes": "IG: @emotors.vcp",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Nicho innovador: motos eléctricas. Sin web. E-commerce ideal para venta de motos eléctricas + repuestos online.",
    },
    {
        "nombre": "MOTOCARCANO",
        "ubicacion": "Villa Carlos Paz",
        "direccion": "Av. Perón 725, Villa Carlos Paz",
        "telefono": "Desde Portal Moto Latino",
        "tipo": "Motos + cuatriciclos",
        "url_web": "Sin web propia",
        "redes": "Listing en Portal Moto Latino",
        "estado_web": "Sin web (solo marketplace)",
        "prioridad": "Alta",
        "notas": "Solo presencia en marketplace de terceros. Sin web propia ni redes activas. Pitch directo para Landing Page.",
    },

    # ===== LA FALDA =====
    {
        "nombre": "La Falda Automotores",
        "ubicacion": "La Falda",
        "direccion": "Av. España 1273, La Falda",
        "telefono": "03548-422170 · +5493548554319",
        "tipo": "Autos 0km + usados multimarca",
        "url_web": "https://lafaldaautomotores.com",
        "redes": "IG: @lafaldaautomotores · 35 años trayectoria",
        "estado_web": "Existe (bloquea bots — verificar)",
        "prioridad": "Media",
        "notas": "35 años trayectoria, +100 vehículos. Sitio web dio 403 a bots (puede estar bloqueado o mal configurado). Revisar manualmente — oportunidad de refresh.",
    },
    {
        "nombre": "SP Automotores",
        "ubicacion": "La Falda",
        "direccion": "Av. España 1186, La Falda",
        "telefono": "Desde Canva site",
        "tipo": "Vehículos exclusivos",
        "url_web": "https://spautomotores.my.canva.site",
        "redes": "—",
        "estado_web": "Amateur (Canva)",
        "prioridad": "Alta",
        "notas": "Hecho en Canva — nada profesional. Sin SEO. Oportunidad clarísima de Landing Page.",
    },
    {
        "nombre": "Paolini Automotores",
        "ubicacion": "La Falda",
        "direccion": "Av. España 601, La Falda",
        "telefono": "WhatsApp 3548 468411 · 3548 468710",
        "tipo": "Renault 0km (Sandero) + usados",
        "url_web": "Sin web propia",
        "redes": "IG: @paoliniautomotores",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Solo Instagram. Vende Sandero 0km con entrega inmediata. Candidato directo para Landing Page.",
    },
    {
        "nombre": "Ruta 38 Automotores",
        "ubicacion": "La Falda",
        "direccion": "La Falda (sobre Ruta 38)",
        "telefono": "Desde FB",
        "tipo": "Autos usados multimarca",
        "url_web": "Sin web propia",
        "redes": "FB: Ruta 38 Automotores · +20 años experiencia",
        "estado_web": "Sin web (solo Facebook)",
        "prioridad": "Alta",
        "notas": "+20 años experiencia. Solo FB. Nombre sugiere ubicación sobre ruta 38 — alto tráfico. Landing Page con catálogo de usados.",
    },
    {
        "nombre": "Dos Ruedas La Falda",
        "ubicacion": "La Falda",
        "direccion": "Av. Buenos Aires 565, La Falda",
        "telefono": "Desde IG/Wix",
        "tipo": "Motos + cuatriciclos + service técnico",
        "url_web": "https://dosruedasconsultas.wixsite.com/dos-ruedas",
        "redes": "IG: @dosruedas_lafalda",
        "estado_web": "Amateur (Wix)",
        "prioridad": "Alta",
        "notas": "Sitio hecho en Wix — limitado y amateur. Servicio técnico de motos y cuatris. Oportunidad de refresh a sitio profesional con catálogo.",
    },

    # ===== VALLE HERMOSO =====
    {
        "nombre": "Automotores Valle Hermoso",
        "ubicacion": "Valle Hermoso",
        "direccion": "Gral. Paz 32, Valle Hermoso",
        "telefono": "3513568287 (desde IG)",
        "tipo": "Autos usados multimarca",
        "url_web": "Sin web propia",
        "redes": "IG + listing gtm.com.ar",
        "estado_web": "Sin web (solo IG + listing)",
        "prioridad": "Alta",
        "notas": "Pueblo chico, sin web propia. Aparece en Instagram y listing de terceros. Oportunidad clara — competencia local casi nula.",
    },
    {
        "nombre": "Baco Automotores",
        "ubicacion": "Valle Hermoso",
        "direccion": "Ruta Nacional 38 N° 299, Valle Hermoso",
        "telefono": "Desde FB",
        "tipo": "Autos usados + créditos prendarios",
        "url_web": "Sin web propia",
        "redes": "FB: Baco Automotores",
        "estado_web": "Sin web (solo Facebook)",
        "prioridad": "Alta",
        "notas": "Sobre Ruta 38 — alta exposición. Solo Facebook. Créditos prendarios en pesos. Landing Page con calculadora de cuotas sería diferencial.",
    },

    # ===== COSQUÍN =====
    {
        "nombre": "Automotores Chahin",
        "ubicacion": "Cosquín",
        "direccion": "Av. San Martín 361, Cosquín",
        "telefono": "(0351) 4657536 · (03543) ... (suc. Cosquín)",
        "tipo": "0km oficial KIA + Mitsubishi",
        "url_web": "https://www.automotoreschahin.com.ar",
        "redes": "IG: @automotoreschahin",
        "estado_web": "Obsoleta (no responsive)",
        "prioridad": "Media",
        "notas": "Concesionario oficial KIA y Mitsubishi. Web obsoleta — sin viewport, no responsive, solo 900 bytes. Marca oficial debería exigir mejor imagen. Oportunidad alta de refresh.",
    },
    {
        "nombre": "Radke Automotores",
        "ubicacion": "Cosquín",
        "direccion": "San Martín esq. Libertad, Cosquín",
        "telefono": "3541576579",
        "tipo": "Autos usados (remolques también)",
        "url_web": "Sin web propia",
        "redes": "FB: Radke Automotores",
        "estado_web": "Sin web (solo Facebook)",
        "prioridad": "Alta",
        "notas": "Solo Facebook. Esquina muy transitada de Cosquín. Pitch directo para Landing Page.",
    },
    {
        "nombre": "Salman Automotores",
        "ubicacion": "Cosquín",
        "direccion": "Av. Cap. Aviador Omar Castillo 2023, Cosquín",
        "telefono": "Desde gtm.com.ar",
        "tipo": "Autos usados",
        "url_web": "Sin web propia",
        "redes": "Listing gtm.com.ar",
        "estado_web": "Sin web (solo listing)",
        "prioridad": "Alta",
        "notas": "Solo listado en guía de concesionarias. Sin web ni redes activas. Oportunidad de Landing Page.",
    },
    {
        "nombre": "J.P. Automotores",
        "ubicacion": "Cosquín",
        "direccion": "Av. A. Sabattini 4290, Cosquín",
        "telefono": "Desde InfoisInfo",
        "tipo": "0km + usados (Creditfacil)",
        "url_web": "Sin web propia",
        "redes": "Listing InfoisInfo",
        "estado_web": "Sin web (solo listing)",
        "prioridad": "Alta",
        "notas": "Solo listing en InfoisInfo. Financiación con Creditfacil. Landing Page con simulador de cuotas.",
    },
    {
        "nombre": "Temprana Automotores",
        "ubicacion": "Cosquín",
        "direccion": "Tucumán 1067, Cosquín",
        "telefono": "(03541) 45-1309",
        "tipo": "Autos usados",
        "url_web": "Sin web propia",
        "redes": "Listing licuo.com.ar",
        "estado_web": "Sin web (solo listing)",
        "prioridad": "Alta",
        "notas": "Solo en guía telefónica. Sin web ni redes. Pitch directo.",
    },
    {
        "nombre": "Cetrogar Motos",
        "ubicacion": "Cosquín",
        "direccion": "Cosquín (dirección no confirmada)",
        "telefono": "Desde web",
        "tipo": "Motos Honda + Yamaha + Corven oficial",
        "url_web": "https://cetrogarmotos.com.ar",
        "redes": "—",
        "estado_web": "Existe (no accesible a bots)",
        "prioridad": "Media",
        "notas": "Concesionario oficial Honda, Yamaha, Corven-Bajaj. Sitio registró timeout — verificar. Cadena Cetrogar es regional fuerte.",
    },
    {
        "nombre": "TodoMoto Cosquín",
        "ubicacion": "Cosquín",
        "direccion": "Cosquín",
        "telefono": "Desde FB",
        "tipo": "Repuestos motos + accesorios",
        "url_web": "Sin web propia",
        "redes": "FB: TodoMoto Cosquin",
        "estado_web": "Sin web (solo Facebook)",
        "prioridad": "Ala",
        "notas": "Rubro repuestos. Solo FB. E-commerce ideal para venta online de repuestos con envío a todo el valle.",
    },

    # ===== CAPILLA DEL MONTE =====
    {
        "nombre": "Rios Automotores",
        "ubicacion": "Capilla del Monte",
        "direccion": "Capilla del Monte",
        "telefono": "3548481809",
        "tipo": "Autos 0km + usados familiar",
        "url_web": "Sin web propia",
        "redes": "FB + IG: @riosautomotores",
        "estado_web": "Sin web (solo FB + IG + listing)",
        "prioridad": "Alta",
        "notas": "Concesionaria familiar. Sin web propia. Triple presencia social pero ningún sitio. Pitch cálido — tono familiar.",
    },
    {
        "nombre": "Sus Autos",
        "ubicacion": "Capilla del Monte",
        "direccion": "38 y Pio Collivadino, Capilla del Monte",
        "telefono": "3548-404520",
        "tipo": "Autos + motos 0km y usados",
        "url_web": "Sin web propia",
        "redes": "IG: @susautosok",
        "estado_web": "Sin web (solo Instagram)",
        "prioridad": "Alta",
        "notas": "Solo Instagram. Mixto autos + motos. Landing Page con dos secciones (autos/motos).",
    },

    # ===== LA CUMBRE =====
    {
        "nombre": "Reyna Automotores",
        "ubicacion": "La Cumbre",
        "direccion": "25 de Mayo 448, La Cumbre",
        "telefono": "(03548) 452-107",
        "tipo": "Autos usados",
        "url_web": "Sin web propia",
        "redes": "Listing gtm.com.ar + licuo.com.ar",
        "estado_web": "Sin web (solo listings)",
        "prioridad": "Alta",
        "notas": "Solo en guías telefónicas. Sin web ni redes. Pueblo chico — poca competencia. Pitch directo.",
    },
]

# Fix small typo
for d in DEALERSHIPS:
    if d["prioridad"] == "Ala":
        d["prioridad"] = "Alta"


# ============================================================
# CREATE WORKBOOK
# ============================================================
wb = Workbook()
wb.properties.creator = "Z.ai · Paulero Studio"

# ===== Sheet 1: Relevamiento completo =====
ws = wb.active
ws.title = "Relevamiento"

HEADERS = [
    "#", "Nombre", "Ubicación", "Dirección", "Teléfono",
    "Tipo", "Sitio web", "Redes sociales",
    "Estado del sitio", "Prioridad", "Notas / Pitch sugerido"
]
COL_WIDTHS = [4, 28, 18, 38, 28, 32, 38, 30, 30, 12, 60]

# Title row
ws.merge_cells("A1:K1")
title_cell = ws["A1"]
title_cell.value = "Relevamiento de concesionarias — Valle de Punilla"
title_cell.font = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor=PRIMARY)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# Subtitle
ws.merge_cells("A2:K2")
sub_cell = ws["A2"]
sub_cell.value = (
    "Paulero Studio · Villa Carlos Paz · La Falda · Valle Hermoso · Cosquín · Capilla del Monte · La Cumbre · "
    f"Total: {len(DEALERSHIPS)} concesionarias relevadas"
)
sub_cell.font = Font(name=FONT_NAME, size=10, italic=True, color=NEUTRAL_600)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 22

# Header row (row 3)
HEADER_ROW = 3
for i, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEADER_ROW, column=i, value=h)
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=NEUTRAL_200),
        right=Side(style="thin", color=NEUTRAL_200),
        bottom=Side(style="medium", color=PRIMARY),
    )
ws.row_dimensions[HEADER_ROW].height = 36

# Column widths
for i, w in enumerate(COL_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Sort by priority (Alta first) then by location
PRIORITY_ORDER = {"Alta": 0, "Media": 1, "Baja": 2}
DEALERSHIPS_SORTED = sorted(DEALERSHIPS, key=lambda d: (PRIORITY_ORDER.get(d["prioridad"], 3), d["ubicacion"], d["nombre"]))

# Data rows
thin_border = Border(
    left=Side(style="thin", color=NEUTRAL_200),
    right=Side(style="thin", color=NEUTRAL_200),
    top=Side(style="thin", color=NEUTRAL_200),
    bottom=Side(style="thin", color=NEUTRAL_200),
)

for idx, d in enumerate(DEALERSHIPS_SORTED, start=1):
    row = HEADER_ROW + idx
    values = [
        idx,
        d["nombre"],
        d["ubicacion"],
        d["direccion"],
        d["telefono"],
        d["tipo"],
        d["url_web"],
        d["redes"],
        d["estado_web"],
        d["prioridad"],
        d["notas"],
    ]
    # Alternating row fill
    row_fill = PatternFill("solid", fgColor=NEUTRAL_100) if idx % 2 == 0 else PatternFill("solid", fgColor=NEUTRAL_0)
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name=FONT_NAME, size=10, color=NEUTRAL_900)
        cell.fill = row_fill
        cell.border = thin_border
        # Alignment per column
        if col == 1:  # #
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 10:  # Prioridad
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
        elif col in (2, 3):  # Nombre, Ubicación
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.font = Font(name=FONT_NAME, size=10, bold=True, color=NEUTRAL_900)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    # Set row height
    ws.row_dimensions[row].height = 60

# Conditional formatting on Prioridad column (J = col 10)
last_row = HEADER_ROW + len(DEALERSHIPS_SORTED)
prio_range = f"J{HEADER_ROW+1}:J{last_row}"

# Alta = red bg
ws.conditional_formatting.add(
    prio_range,
    CellIsRule(operator="equal", formula=['"Alta"'],
               fill=PatternFill("solid", fgColor="FDEDEC"),
               font=Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_NEGATIVE))
)
# Media = amber bg
ws.conditional_formatting.add(
    prio_range,
    CellIsRule(operator="equal", formula=['"Media"'],
               fill=PatternFill("solid", fgColor="FEF9E7"),
               font=Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_WARNING))
)
# Baja = green bg
ws.conditional_formatting.add(
    prio_range,
    CellIsRule(operator="equal", formula=['"Baja"'],
               fill=PatternFill("solid", fgColor="E8F5E9"),
               font=Font(name=FONT_NAME, size=10, bold=True, color=ACCENT_POSITIVE))
)

# Freeze panes (header + # column)
ws.freeze_panes = "C4"

# Auto filter
ws.auto_filter.ref = f"A{HEADER_ROW}:K{last_row}"

# ============================================================
# Sheet 2: Resumen por ubicación
# ============================================================
ws2 = wb.create_sheet("Resumen por zona")

# Count per location and priority
from collections import Counter
loc_counts = Counter(d["ubicacion"] for d in DEALERSHIPS)
loc_prio = {}
for d in DEALERSHIPS:
    loc_prio.setdefault(d["ubicacion"], {"Alta": 0, "Media": 0, "Baja": 0})
    loc_prio[d["ubicacion"]][d["prioridad"]] += 1

# Sort locations by total desc
sorted_locs = sorted(loc_counts.items(), key=lambda x: -x[1])

# Title
ws2.merge_cells("A1:F1")
t = ws2["A1"]
t.value = "Resumen por zona — Valle de Punilla"
t.font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=PRIMARY)
t.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Header
HEADERS2 = ["Ubicación", "Total", "Prioridad Alta 🔴", "Prioridad Media 🟡", "Prioridad Baja 🟢", "% Sin web / Amateur"]
for i, h in enumerate(HEADERS2, start=1):
    c = ws2.cell(row=3, column=i, value=h)
    c.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border
ws2.row_dimensions[3].height = 32

# Column widths
for i, w in enumerate([22, 10, 18, 18, 18, 22], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Data
for i, (loc, total) in enumerate(sorted_locs, start=1):
    row = 3 + i
    p = loc_prio[loc]
    no_web_pct = round((p["Alta"]) / total * 100, 0)
    vals = [loc, total, p["Alta"], p["Media"], p["Baja"], f"{int(no_web_pct)}%"]
    row_fill = PatternFill("solid", fgColor=NEUTRAL_100) if i % 2 == 0 else PatternFill("solid", fgColor=NEUTRAL_0)
    for col, v in enumerate(vals, start=1):
        c = ws2.cell(row=row, column=col, value=v)
        c.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)
        c.fill = row_fill
        c.border = thin_border
        if col == 1:
            c.font = Font(name=FONT_NAME, size=11, bold=True, color=NEUTRAL_900)
            c.alignment = Alignment(horizontal="left", vertical="center")
        elif col == 6:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name=FONT_NAME, size=11, bold=True, color=ACCENT_NEGATIVE if no_web_pct >= 60 else ACCENT_WARNING if no_web_pct >= 30 else ACCENT_POSITIVE)
        else:
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[row].height = 28

# Totals row
total_row = 3 + len(sorted_locs) + 1
ws2.cell(row=total_row, column=1, value="TOTAL Valle de Punilla").font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
ws2.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=PRIMARY)
ws2.cell(row=total_row, column=1).alignment = Alignment(horizontal="right", vertical="center")
ws2.cell(row=total_row, column=1).border = thin_border
for col, val in enumerate([
    len(DEALERSHIPS),
    sum(1 for d in DEALERSHIPS if d["prioridad"] == "Alta"),
    sum(1 for d in DEALERSHIPS if d["prioridad"] == "Media"),
    sum(1 for d in DEALERSHIPS if d["prioridad"] == "Baja"),
    f"{round(sum(1 for d in DEALERSHIPS if d['prioridad'] == 'Alta') / len(DEALERSHIPS) * 100)}%"
], start=2):
    c = ws2.cell(row=total_row, column=col, value=val)
    c.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = thin_border
ws2.row_dimensions[total_row].height = 30

# Notes row
note_row = total_row + 2
ws2.merge_cells(f"A{note_row}:F{note_row}")
nc = ws2.cell(row=note_row, column=1)
nc.value = (
    "🔴 Prioridad Alta = Sin web propia / Solo redes sociales / Web amateur (Canva, Wix, Google Sites). "
    "🟡 Prioridad Media = Web obsoleta (no responsive) o inaccesible a verificación. "
    "🟢 Prioridad Baja = Web decente (probablemente ya inviertan en imagen). "
    "Recomendación: arrancar por las Altas — son las que más necesitan y mejor cierran."
)
nc.font = Font(name=FONT_NAME, size=10, italic=True, color=NEUTRAL_600)
nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[note_row].height = 60

# ============================================================
# Sheet 3: Cómo usar esta planilla
# ============================================================
ws3 = wb.create_sheet("Cómo usar")

ws3.column_dimensions["A"].width = 4
ws3.column_dimensions["B"].width = 100

ws3.merge_cells("A1:B1")
t = ws3["A1"]
t.value = "Cómo usar esta planilla"
t.font = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
t.fill = PatternFill("solid", fgColor=PRIMARY)
t.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 32

instructions = [
    ("1.", "Arrancá por las prioridad ALTA — son concesionarias sin web o con web amateur (Canva/Wix/Google Sites). Esas son las que más rápido cierran porque el dolor es evidente."),
    ("2.", "Para cada una, mandá WhatsApp con un pitch directo: 'Hola [Nombre], vi que tienen concesionaria en [pueblo] y noté que no tienen sitio web propio / el que tienen está hecho en Canva. Hago webs profesionales a medida desde 250 USD pago único. ¿Tenés 5 minutos para charlar?'"),
    ("3.", "Mostrá tu portfolio (paulerostudio.com) y mencioná que ya trabajaste con otros negocios del rubro (Compucity, Etersomos)."),
    ("4.", "Para los de prioridad MEDIA (web obsoleta o inaccesible), el pitch es de REFRESH: 'Tu web actual no se ve bien en celular y Google no la indexa bien. Por X USD te la renuevo completa con diseño a medida.'"),
    ("5.", "Para los de prioridad BAJA (web decente), no insistas con sitio web. Vendéles CHATBOT IA: 'Tu web está bien, pero ¿te atiende 24/7? Te puedo integrar un chatbot con IA que califica leads y deriva a WhatsApp.'"),
    ("6.", "Usá el código STUDIO20 del chatbot para tracking. Si mencionan que vienen del chatbot de tu web, aplicá 20% off."),
    ("7.", "Trackeá en esta planilla: cuando hables con uno, agregá una columna 'Estado contacto' y 'Próxima acción'."),
    ("8.", "Zonas con mayor concentración de oportunidad (prioridad Alta): Villa Carlos Paz (15 concesionarias) y Cosquín (6). Empezá por ahí."),
]

for i, (n, text) in enumerate(instructions, start=3):
    ws3.cell(row=i, column=1, value=n).font = Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY)
    ws3.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ws3.cell(row=i, column=2, value=text).font = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)
    ws3.cell(row=i, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws3.row_dimensions[i].height = 50

# Save
output_path = "/home/z/my-project/download/Relevamiento_Concesionarias_ValleDePunilla.xlsx"
wb.save(output_path)
print(f"✓ Excel guardado en: {output_path}")
print(f"  - Hoja 1: Relevamiento ({len(DEALERSHIPS)} concesionarias)")
print(f"  - Hoja 2: Resumen por zona")
print(f"  - Hoja 3: Cómo usar")
print()
print("=== Distribución por prioridad ===")
prio_count = Counter(d["prioridad"] for d in DEALERSHIPS)
for p in ["Alta", "Media", "Baja"]:
    print(f"  {p}: {prio_count[p]}")
print()
print("=== Distribución por zona ===")
for loc, cnt in sorted_locs:
    print(f"  {loc}: {cnt}")
